import requests
import re
import openpyxl
from time import sleep
import datetime

day = datetime.datetime.now().day
month = datetime.datetime.now().month
year = datetime.datetime.now().year

filename = "export " + str(day) + '-' + str(month) + '-' + str(year) + ".xlsx"

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "MST"
ws['B1'] = "TÊN CÔNG TY"
ws['C1'] = "ĐỊA CHỈ"
ws['D1'] = "ĐẠI DIỆN PHÁP LUẬT"
ws['E1'] = "ĐIỆN THOẠI"
ws['F1'] = "NGÀY HOẠT ĐỘNG"
ws['G1'] = "TÌNH TRẠNG"
ws['H1'] = "HOÁ ĐƠN ĐIỆN TỬ"
wb.save(filename)

def update_value_excel(filename,cellname,value):
    wb = openpyxl.load_workbook(filename)
    Sheet = wb['Sheet']
    Sheet[cellname] = value
    wb.save(filename)


listcode = open("taxcodelist.txt", "r").read()
array = listcode.split()
arrayLength = str(len(array))


j = 1
for i in array:
    sleep(1)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:96.0) Gecko/20100101 Firefox/96.0"})
    html = session.get("https://masothue.com/Search/?q=" + i + "&type=enterpriseTax").text
    check = html.find('table class="table-taxinfo"')
    if check < 0:
        start = html.find('class="tax-listing"') + len('class="tax-listing"')
        end = html.find('id="sidebar"')
        table = html[start:end]
        bs1 = table.find('<div data-prefetch')
        be1 = table.find('</address>')
        block = table[bs1:be1]
        # print(block)
        cns1 = block.find("href='/")
        cne1 = block.find("' title")
        cn1 = block[cns1:cne1].replace("href='/","")
        link = 'https://masothue.vn/' + cn1
        session = requests.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:96.0) Gecko/20100101 Firefox/96.0"})
        html = session.get(link).text
        start = html.find('table class="table-taxinfo"') + len('table class="table-taxinfo"')
        end = html.find("</table>")
        table = html[start:end]
        # GET COMPANY NAME
        cns1 = table.find('colspan="2"')
        cne1 = table.find('</thead>')
        cn1 = table[cns1:cne1]
        cns2 = cn1.find('copy">')
        cne2 = cn1.find('</span')
        companyname = cn1[cns2:cne2].replace('copy">','')
        print('---' + str(j) + '/' + arrayLength + '---')
        print(i)
        print(companyname)
        print('')
        # GET ADDRESS
        as1 = table.find('address')
        ae1 = table.find('alumni')
        a = table[as1:ae1]
        as2 = a.find("copy'>")
        ae2 = a.find('</span>')
        address = a[as2:ae2].replace("copy'>","")
        # GET DIRECTOR
        ds1 = table.find('legalName')
        de1 = table.find('fa-phone')
        d = table[ds1:de1]
        ds2 = d.find("'>")
        de2 = d.find('</a>')
        director = d[ds2:de2].replace("'>","")
        # GET PHONE
        ps1 = table.find('telephone')
        pe2 = table.find('fa-calendar')
        p1 = table[ps1:pe2]
        ps2 = p1.find(">0")
        pe2 = p1.find("</span")
        p2 = p1[ps2:pe2]
        phone = p2.replace('>','')
        # GET ACTIVE
        acs1 = table.find('calendar')
        ace2 = table.find('users')
        ac1 = table[acs1:ace2]
        acs2 = ac1.find("copy'>")
        ace2 = ac1.find('</span>')
        ac2 = ac1[acs2:ace2]
        active = ac2.replace("copy'>","")
        # GET STATUS
        stts1 = table.find('info')
        stte2 = table.find('<em>')
        stt1 = table[stts1:stte2]
        stts2 = stt1.find(")'>")
        stte2 = stt1.find('</a>')
        stt2 = stt1[stts2:stte2]
        status = stt2.replace(")'>","")
        # GET INVOICE
        invs1 = html.find('table-taxinfo')
        inve2 = html.find('3079208226')
        inv1 = html[invs1:inve2]
        invs2 = inv1.find("_blank'>")
        inve2 = inv1.find('async')
        inv2 = inv1[invs2:inve2]
        invs3 = inv2.find("k'>")
        inve3 = inv2.find("</a>")
        inv3 = inv2[invs3:inve3]
        invoice = inv3.replace("k'>","")
        j = j + 1
        cell_name_mst = "A" + str(j)
        cell_name_cty = "B" + str(j)
        cell_name_address = "C" + str(j)
        cell_name_director = "D" + str(j)
        cell_name_phone = "E" + str(j)
        cell_name_active = "F" + str(j)
        cell_name_status = "G" + str(j)
        cell_name_invoice = "H" + str(j)
        update_value_excel(filename,cell_name_mst,str(i))
        update_value_excel(filename,cell_name_cty,companyname)
        update_value_excel(filename,cell_name_address,address)
        update_value_excel(filename,cell_name_director,director)
        update_value_excel(filename,cell_name_phone,phone)
        update_value_excel(filename,cell_name_active,active)
        update_value_excel(filename,cell_name_status,status)
        update_value_excel(filename,cell_name_invoice,invoice)
    else:
        start = html.find('table class="table-taxinfo"') + len('table class="table-taxinfo"')
        end = html.find("</table>")
        table = html[start:end]
        # GET COMPANY NAME
        cns1 = table.find('colspan="2"')
        cne1 = table.find('</thead>')
        cn1 = table[cns1:cne1]
        cns2 = cn1.find('copy">')
        cne2 = cn1.find('</span')
        companyname = cn1[cns2:cne2].replace('copy">','')
        print('---' + str(j) + '/' + arrayLength + '---')
        print(i)
        print(companyname)
        print('')
        # GET ADDRESS
        as1 = table.find('address')
        ae1 = table.find('alumni')
        a = table[as1:ae1]
        as2 = a.find("copy'>")
        ae2 = a.find('</span>')
        address = a[as2:ae2].replace("copy'>","")
        # GET DIRECTOR
        ds1 = table.find('legalName')
        de1 = table.find('fa-phone')
        d = table[ds1:de1]
        ds2 = d.find("'>")
        de2 = d.find('</a>')
        director = d[ds2:de2].replace("'>","")
        # GET PHONE
        ps1 = table.find('telephone')
        pe2 = table.find('fa-calendar')
        p1 = table[ps1:pe2]
        ps2 = p1.find(">0")
        pe2 = p1.find("</span")
        p2 = p1[ps2:pe2]
        phone = p2.replace('>','')
        # GET ACTIVE
        acs1 = table.find('calendar')
        ace2 = table.find('users')
        ac1 = table[acs1:ace2]
        acs2 = ac1.find("copy'>")
        ace2 = ac1.find('</span>')
        ac2 = ac1[acs2:ace2]
        active = ac2.replace("copy'>","")
        # GET STATUS
        stts1 = table.find('info')
        stte2 = table.find('<em>')
        stt1 = table[stts1:stte2]
        stts2 = stt1.find(")'>")
        stte2 = stt1.find('</a>')
        stt2 = stt1[stts2:stte2]
        status = stt2.replace(")'>","")
        # GET INVOICE
        invs1 = html.find('table-taxinfo')
        inve2 = html.find('3079208226')
        inv1 = html[invs1:inve2]
        invs2 = inv1.find("_blank'>")
        inve2 = inv1.find('async')
        inv2 = inv1[invs2:inve2]
        invs3 = inv2.find("k'>")
        inve3 = inv2.find("</a>")
        inv3 = inv2[invs3:inve3]
        invoice = inv3.replace("k'>","")
        j = j + 1
        cell_name_mst = "A" + str(j)
        cell_name_cty = "B" + str(j)
        cell_name_address = "C" + str(j)
        cell_name_director = "D" + str(j)
        cell_name_phone = "E" + str(j)
        cell_name_active = "F" + str(j)
        cell_name_status = "G" + str(j)
        cell_name_invoice = "H" + str(j)
        update_value_excel(filename,cell_name_mst,str(i))
        update_value_excel(filename,cell_name_cty,companyname)
        update_value_excel(filename,cell_name_address,address)
        update_value_excel(filename,cell_name_director,director)
        update_value_excel(filename,cell_name_phone,phone)
        update_value_excel(filename,cell_name_active,active)
        update_value_excel(filename,cell_name_status,status)
        update_value_excel(filename,cell_name_invoice,invoice)


def close_excel(filename):
    wb = openpyxl.load_workbook(filename)
    wb.close()
    wb.save(filename)







